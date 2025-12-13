from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.page import PageMargins
from PIL import Image as PILImage
from datetime import datetime
import os
import re
import subprocess
import shutil
import tempfile
import uuid

# -------------------------------------------------
# YOLLAR
# -------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_FILE = os.path.join(BASE_DIR, "rapor.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "generated_pdfs")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# -------------------------------------------------
# FOTOĞRAF ALANLARI
# -------------------------------------------------
PHOTO_AREAS = {
    1: ("B", 29, "I", 47),
    2: ("J", 29, "P", 47),
    3: ("B", 52, "I", 70),
    4: ("J", 52, "P", 70),
    5: ("B", 75, "I", 93),
    6: ("J", 75, "P", 93),
    7: ("B", 98, "I", 116),
    8: ("J", 98, "P", 116),
}

# -------------------------------------------------
# TÜRKÇE NORMALIZE
# -------------------------------------------------
def normalize(text: str) -> str:
    return (
        text.lower()
        .replace("ı", "i")
        .replace("ğ", "g")
        .replace("ü", "u")
        .replace("ş", "s")
        .replace("ö", "o")
        .replace("ç", "c")
        .replace(":", "")
        .strip()
    )

# -------------------------------------------------
# FOTOĞRAF ALAN BOYUTU (PX)
# -------------------------------------------------
def area_pixel_size(ws, area):
    sc, sr, ec, er = area

    width = sum(
        (ws.column_dimensions[ws.cell(1, c).column_letter].width or 8) * 7
        for c in range(column_index_from_string(sc), column_index_from_string(ec) + 1)
    )

    height = sum(
        (ws.row_dimensions[r].height or 15) * 1.25
        for r in range(sr, er + 1)
    )

    return int(width * 0.95), int(height * 0.95)

# -------------------------------------------------
# GERÇEK MERKEZ ANCHOR
# -------------------------------------------------
def smart_anchor(ws, area, img_height_px):
    sc, sr, ec, er = area
    sc_i = column_index_from_string(sc)
    ec_i = column_index_from_string(ec)

    center_col = sc_i + (ec_i - sc_i) // 2
    center_row = sr + (er - sr) // 2

    px_per_row = 18
    row_offset = int((img_height_px / 2) / px_per_row)

    final_row = max(sr + 1, center_row - row_offset)
    col_letter = ws.cell(row=1, column=center_col).column_letter

    return f"{col_letter}{final_row}"

# -------------------------------------------------
# PRINT AREA VE SAYFA AYARLARI
# -------------------------------------------------
def setup_print_area(ws):
    """
    Excel sayfasına print area (yazdırma alanı) ve sayfa ayarları yapar.
    Tüm içeriği ve resimleri PDF'e dahil etmek için.
    """
    # En son dolu satırı bul
    max_row = ws.max_row
    # Resimlerin olduğu alanları kontrol et
    for area in PHOTO_AREAS.values():
        _, _, _, er = area
        if er > max_row:
            max_row = er
    
    # En son dolu sütunu bul
    max_col = ws.max_column
    # Resimlerin olduğu alanları kontrol et
    for area in PHOTO_AREAS.values():
        _, _, ec, _ = area
        ec_i = column_index_from_string(ec)
        if ec_i > max_col:
            max_col = ec_i
    
    # Print area'yı ayarla (A1'den en son dolu hücreye kadar)
    # Son foto alanının altına kadar
    print_end_col = get_column_letter(max_col)
    print_end_row = max_row + 2  # Biraz ekstra alan
    
    print_area = f"A1:{print_end_col}{print_end_row}"
    ws.print_area = print_area
    
    # Sayfa ayarları
    ws.page_setup.orientation = "portrait"  # "portrait" veya "landscape"
    ws.page_setup.paperSize = 9  # 9 = A4
    
    # Scale'i dinamik olarak hesapla - içeriğin yüksekliğine göre
    # A4 sayfa yüksekliği: ~29.7 cm = ~842 point
    # Excel satır yüksekliği ortalama: ~15 point
    # İçerik yüksekliği: max_row * 15 point (yaklaşık)
    
    # İçeriğin toplam yüksekliğini hesapla
    total_height_points = 0
    for row in range(1, max_row + 1):
        row_height = ws.row_dimensions[row].height
        if row_height:
            total_height_points += row_height * 1.25  # Excel'de point to pixel dönüşümü
        else:
            total_height_points += 15 * 1.25  # Varsayılan satır yüksekliği
    
    # A4 sayfa yüksekliği (point cinsinden, kenar boşlukları hariç)
    # A4: 29.7 cm = 842 point, kenar boşlukları: 0.4 inch = ~28 point (üst+alt)
    usable_page_height = 842 - 56  # ~786 point kullanılabilir alan
    
    # Genişliği de hesapla (yana taşmayı önlemek için)
    total_width_points = 0
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        col_width = ws.column_dimensions[col_letter].width
        if col_width:
            total_width_points += col_width * 7  # Excel'de genişlik to point dönüşümü
        else:
            total_width_points += 8 * 7  # Varsayılan sütun genişliği
    
    # A4 sayfa genişliği (point cinsinden, kenar boşlukları hariç)
    # A4: 21 cm = 595 point, kenar boşlukları: 0.4 inch = ~28 point (sol+sağ)
    usable_page_width = 595 - 56  # ~539 point kullanılabilir alan
    
    # Scale hesapla: hem yüksekliğe hem genişliğe göre
    scale_by_height = 100
    scale_by_width = 100
    
    if total_height_points > 0:
        scale_by_height = int((usable_page_height / total_height_points) * 100)
    
    if total_width_points > 0:
        scale_by_width = int((usable_page_width / total_width_points) * 100)
    
    # İkisinden küçük olanı al (hem yüksekliğe hem genişliğe sığsın)
    calculated_scale = min(scale_by_height, scale_by_width)
    
    # Biraz daha küçült (güvenlik için %10 daha küçük)
    calculated_scale = int(calculated_scale * 0.90)
    
    # Scale'i %40-%85 arasında tut (çok küçük veya çok büyük olmasın)
    calculated_scale = max(40, min(85, calculated_scale))
    
    # Scale'i ayarla
    try:
        ws.page_setup.scale = calculated_scale
    except Exception:
        # Scale ayarlanamazsa varsayılan değer
        try:
            ws.page_setup.scale = 60
        except Exception:
            pass
    
    # fitToPage'i kaldır - sadece scale kullan
    try:
        if hasattr(ws, 'sheet_properties') and ws.sheet_properties:
            if hasattr(ws.sheet_properties, 'pageSetUpPr'):
                if ws.sheet_properties.pageSetUpPr is not None:
                    ws.sheet_properties.pageSetUpPr.fitToPage = False
                    ws.sheet_properties.pageSetUpPr.fitToWidth = 0
                    ws.sheet_properties.pageSetUpPr.fitToHeight = 0
    except Exception:
        pass
    
    # Kenar boşlukları (margins) - tek sayfaya sığdırmak için küçük
    try:
        ws.page_margins = PageMargins(
            left=0.2,
            right=0.2,
            top=0.2,
            bottom=0.2,
            header=0.1,
            footer=0.1
        )
    except Exception:
        pass  # Bazı şablonlarda çalışmayabilir, devam et
    
    return print_area

# -------------------------------------------------
# EXCEL → PDF (xlsx2pdf - önerilen)
# -------------------------------------------------
def excel_to_pdf_xlsx2pdf(excel_file, pdf_file):
    """
    Excel dosyasını PDF'e çevirir.
    xlsx2pdf 1.0.4 kullanır - Transformer sınıfı ile.
    """
    try:
        # xlsx2pdf 1.0.4'te Transformer sınıfı kullanılır
        from xlsx2pdf.transformator import Transformer
        from openpyxl import load_workbook
        
        print(f"Transformer ile PDF oluşturuluyor: {excel_file} -> {pdf_file}")
        
        # Transformer'ın __init__ metodunda font sorunu var
        # Workbook objesi ile deneyelim
        wb = load_workbook(excel_file)
        transformer = Transformer(wb)
        
        # transform() veya convert() metodunu dene
        if hasattr(transformer, 'transform'):
            transformer.transform(pdf_file)
        elif hasattr(transformer, 'convert'):
            transformer.convert(pdf_file)
        elif hasattr(transformer, 'to_pdf'):
            transformer.to_pdf(pdf_file)
        else:
            # Transformer'ın metodlarını listele
            methods = [m for m in dir(transformer) if not m.startswith('_') and callable(getattr(transformer, m))]
            print(f"Transformer metodları: {methods}")
            raise Exception(f"Transformer'da transform/convert/to_pdf metodu bulunamadı. Mevcut metodlar: {methods}")
        
        if os.path.exists(pdf_file) and os.path.getsize(pdf_file) > 0:
            print(f"PDF başarıyla oluşturuldu: {pdf_file}, boyut: {os.path.getsize(pdf_file)} bytes")
            return True
        else:
            print("PDF dosyası oluşturulamadı veya boş")
            return False
            
    except (TypeError, AttributeError) as te:
        # Transformer workbook ile çalışmıyorsa, dosya yolu ile dene
        print(f"Transformer workbook ile çalışmadı, dosya yolu ile deneniyor: {te}")
        try:
            from xlsx2pdf.transformator import Transformer
            transformer = Transformer(excel_file)
            if hasattr(transformer, 'transform'):
                transformer.transform(pdf_file)
            elif hasattr(transformer, 'convert'):
                transformer.convert(pdf_file)
            else:
                raise Exception("Transformer'da transform/convert metodu bulunamadı")
            
            if os.path.exists(pdf_file) and os.path.getsize(pdf_file) > 0:
                return True
        except Exception as e2:
            print(f"Dosya yolu ile de çalışmadı: {e2}")
            return False
    except ImportError as e:
        print(f"xlsx2pdf ImportError: {e}")
        return False
    except Exception as e:
        print(f"xlsx2pdf Exception: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return False
        
        print(f"xlsx2pdf modülü yüklendi: {xlsx2pdf}")
        print(f"xlsx2pdf __path__: {getattr(xlsx2pdf, '__path__', 'N/A')}")
        
        # Paket içindeki tüm modülleri listele
        modules = []
        if hasattr(xlsx2pdf, '__path__'):
            for importer, modname, ispkg in pkgutil.iter_modules(xlsx2pdf.__path__):
                modules.append(modname)
                print(f"xlsx2pdf alt modülü bulundu: {modname} (paket: {ispkg})")
        
        # Farklı import yollarını dene
        convert_func = None
        
        # 1. Bulunan modülleri kontrol et (cli, transformator)
        for mod_name in modules + ['converter', 'xlsx2pdf', 'main', 'core']:
            try:
                full_mod_name = f'xlsx2pdf.{mod_name}'
                mod = importlib.import_module(full_mod_name)
                print(f"{full_mod_name} modülü yüklendi: {dir(mod)}")
                
                # Modül içindeki tüm fonksiyonları kontrol et
                for attr_name in dir(mod):
                    if not attr_name.startswith('_'):
                        attr = getattr(mod, attr_name)
                        if callable(attr):
                            print(f"  - {attr_name}: {attr}")
                            # convert, xlsx2pdf, transform gibi isimleri kontrol et
                            if 'convert' in attr_name.lower() or 'transform' in attr_name.lower() or 'xlsx2pdf' in attr_name.lower():
                                convert_func = attr
                                print(f"  -> Potansiyel convert fonksiyonu bulundu: {attr_name}")
                                break
                
                if convert_func:
                    break
                    
                # Direkt convert veya xlsx2pdf fonksiyonlarını kontrol et
                if hasattr(mod, 'convert'):
                    convert_func = getattr(mod, 'convert')
                    print(f"convert fonksiyonu {full_mod_name} modülünde bulundu")
                    break
                elif hasattr(mod, 'xlsx2pdf'):
                    convert_func = getattr(mod, 'xlsx2pdf')
                    print(f"xlsx2pdf fonksiyonu {full_mod_name} modülünde bulundu")
                    break
                elif hasattr(mod, 'transform'):
                    convert_func = getattr(mod, 'transform')
                    print(f"transform fonksiyonu {full_mod_name} modülünde bulundu")
                    break
            except ImportError as e:
                print(f"{full_mod_name} import edilemedi: {e}")
                continue
        
        # 2. cli modülünü subprocess ile çağır (command-line aracı olabilir)
        if convert_func is None:
            try:
                import subprocess
                import sys
                
                # xlsx2pdf cli modülünü Python script olarak çalıştır
                cli_script = f"""
import sys
sys.path.insert(0, '{os.path.dirname(xlsx2pdf.__file__)}')
from xlsx2pdf.cli import main
sys.argv = ['xlsx2pdf', '{excel_file}', '{pdf_file}']
main()
"""
                result = subprocess.run(
                    [sys.executable, '-c', cli_script],
                    capture_output=True,
                    text=True,
                    timeout=60
                )
                if result.returncode == 0 and os.path.exists(pdf_file):
                    print("xlsx2pdf cli modülü ile başarılı")
                    if os.path.getsize(pdf_file) > 0:
                        return True
                else:
                    print(f"xlsx2pdf cli hatası (returncode: {result.returncode}): {result.stderr}")
                    
                # Alternatif: transformator modülünü dene
                try:
                    from xlsx2pdf import transformator
                    print(f"transformator modülü: {dir(transformator)}")
                    # transformator modülünün fonksiyonlarını kontrol et
                    for attr_name in dir(transformator):
                        if not attr_name.startswith('_') and callable(getattr(transformator, attr_name)):
                            print(f"transformator.{attr_name} bulundu")
                            if 'transform' in attr_name.lower() or 'convert' in attr_name.lower():
                                transform_func = getattr(transformator, attr_name)
                                # transformator'un nasıl kullanıldığını dene
                                try:
                                    # Muhtemelen (excel_file, pdf_file) parametreleri alır
                                    transform_func(excel_file, pdf_file)
                                    if os.path.exists(pdf_file) and os.path.getsize(pdf_file) > 0:
                                        print("transformator ile başarılı")
                                        return True
                                except Exception as tf_e:
                                    print(f"transformator.{attr_name} çağrı hatası: {tf_e}")
                except Exception as tr_e:
                    print(f"transformator import/çağrı hatası: {tr_e}")
                    
            except Exception as e:
                print(f"CLI/transformator deneme hatası: {e}")
                import traceback
                traceback.print_exc()
        
        # 3. convert_func bulunduysa kullan
        if convert_func:
            print(f"convert_func çağrılıyor: {convert_func}")
            # Transformer sınıfı ise önce instance oluştur, sonra convert et
            if 'Transformer' in str(type(convert_func)) or 'Transformer' in str(convert_func):
                print("Transformer sınıfı kullanılıyor")
                # Transformer muhtemelen sadece excel_file alıyor (workbook)
                try:
                    transformer = convert_func(excel_file)
                    print(f"Transformer instance oluşturuldu: {type(transformer)}")
                    print(f"Transformer metodları: {[m for m in dir(transformer) if not m.startswith('_') and callable(getattr(transformer, m))]}")
                    
                    # Transformer'ın metodlarını dene
                    if hasattr(transformer, 'convert'):
                        print("Transformer.convert() çağrılıyor")
                        transformer.convert(pdf_file)
                    elif hasattr(transformer, 'to_pdf'):
                        print("Transformer.to_pdf() çağrılıyor")
                        transformer.to_pdf(pdf_file)
                    elif hasattr(transformer, 'save'):
                        print("Transformer.save() çağrılıyor")
                        transformer.save(pdf_file)
                    elif hasattr(transformer, 'write'):
                        print("Transformer.write() çağrılıyor")
                        transformer.write(pdf_file)
                    elif hasattr(transformer, '__call__'):
                        print("Transformer() çağrılıyor (callable)")
                        transformer(pdf_file)
                    else:
                        # Transformer'ın tüm özelliklerini listele
                        all_attrs = [m for m in dir(transformer) if not m.startswith('_')]
                        print(f"Transformer tüm özellikleri: {all_attrs}")
                        raise Exception(f"Transformer'da convert/to_pdf/save/write metodu bulunamadı. Mevcut metodlar: {all_attrs}")
                except TypeError as te:
                    print(f"Transformer oluşturma hatası: {te}")
                    # Belki Transformer farklı parametreler alıyor
                    import inspect
                    try:
                        sig = inspect.signature(convert_func)
                        print(f"Transformer __init__ imzası: {sig}")
                        # Belki workbook objesi gerekiyor
                        from openpyxl import load_workbook
                        wb = load_workbook(excel_file)
                        transformer = convert_func(wb)
                        # Sonra convert et
                        if hasattr(transformer, 'convert'):
                            transformer.convert(pdf_file)
                        elif hasattr(transformer, 'to_pdf'):
                            transformer.to_pdf(pdf_file)
                        elif hasattr(transformer, 'save'):
                            transformer.save(pdf_file)
                        else:
                            raise Exception("Transformer workbook ile oluşturuldu ama convert metodu bulunamadı")
                    except Exception as e2:
                        print(f"Workbook ile deneme hatası: {e2}")
                        raise te
            else:
                # Normal fonksiyon ise direkt çağır
                convert_func(excel_file, pdf_file)
        else:
            raise Exception("xlsx2pdf'te convert fonksiyonu bulunamadı")
        
        # PDF dosyasının oluştuğunu ve boş olmadığını kontrol et
        if os.path.exists(pdf_file):
            file_size = os.path.getsize(pdf_file)
            print(f"PDF dosyası oluşturuldu: {pdf_file}, boyut: {file_size} bytes")
            if file_size > 0:
                return True
            else:
                print("PDF dosyası boş!")
                return False
        else:
            print(f"PDF dosyası oluşturulamadı: {pdf_file}")
            return False
    except ImportError as e:
        print(f"xlsx2pdf ImportError: {e}")
        import traceback
        traceback.print_exc()
        return False
    except Exception as e:
        print(f"xlsx2pdf Exception: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return False

# -------------------------------------------------
# EXCEL → PDF (LibreOffice)
# -------------------------------------------------
def excel_to_pdf_libreoffice(excel_file, pdf_file):
    libreoffice_cmd = (
        shutil.which("soffice")
        or "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    )

    if not libreoffice_cmd or not os.path.exists(libreoffice_cmd):
        return False

    cmd = [
        libreoffice_cmd,
        "--headless",
        "--convert-to", "pdf",
        "--outdir", os.path.dirname(excel_file),
        excel_file
    ]

    subprocess.run(cmd, capture_output=True)

    generated = excel_file.replace(".xlsx", ".pdf")
    if os.path.exists(generated):
        os.replace(generated, pdf_file)
        return True

    return False

# -------------------------------------------------
# METİN PARSE ET
# -------------------------------------------------
def parse_text(text):
    """
    Form'dan gelen metni parse eder.
    Terminal script'indeki mantıkla aynı.
    """
    lines = text.strip().split("\n")
    
    data = {"rapor_no": "", "tarih": "", "yapilan_isler": []}
    mode = None

    for raw in lines:
        clean = normalize(raw)
        if clean == "rapor_no":
            mode = "rapor_no"
            continue
        if clean in ("tarih", "tarih_no"):
            mode = "tarih"
            continue
        if clean in ("yapilan_isler", "yapilan_is"):
            mode = "yapilan_isler"
            continue

        if mode == "rapor_no" and not data["rapor_no"]:
            data["rapor_no"] = raw
        elif mode == "tarih" and not data["tarih"]:
            data["tarih"] = raw
        elif mode == "yapilan_isler" and raw.startswith("-"):
            data["yapilan_isler"].append(raw[1:].strip())

    if not data["tarih"]:
        data["tarih"] = datetime.now().strftime("%d.%m.%Y")

    return data

# -------------------------------------------------
# ANA FONKSİYON
# -------------------------------------------------
def generate_report(text, photos):
    """
    Form'dan gelen metin ve fotoğrafları kullanarak PDF oluşturur.
    
    Args:
        text: Form'dan gelen rapor metni
        photos: Flask FileStorage listesi (fotoğraflar)
    
    Returns:
        PDF dosyasının yolu
    """
    # Metni parse et
    data = parse_text(text)
    safe_date = re.sub(r"[^\d.]", "", data["tarih"])
    
    # Excel şablonunu yükle
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Excel şablonu bulunamadı: {TEMPLATE_FILE}")
    
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    # Excel'i doldur
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Rapor_No":
                cell.value = data["rapor_no"]
            elif cell.value == "Tarih_No":
                cell.value = data["tarih"]
            elif isinstance(cell.value, str) and normalize(cell.value).startswith("foto"):
                cell.value = ""

    # Yapılan işleri ekle
    for i, text_item in enumerate(data["yapilan_isler"]):
        ws[f"B{8 + i}"].value = text_item

    # Fotoğrafları işle
    temp_files = []
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Fotoğrafları geçici dizine kaydet
        photo_files = []
        for i, photo in enumerate(photos[:8]):
            if photo.filename:
                # Dosya uzantısını al
                ext = os.path.splitext(photo.filename)[1] or ".jpg"
                temp_photo_path = os.path.join(temp_dir, f"photo_{i+1}{ext}")
                photo.save(temp_photo_path)
                photo_files.append(temp_photo_path)

        # Fotoğrafları Excel'e ekle
        for i, photo_path in enumerate(photo_files, start=1):
            area = PHOTO_AREAS[i]
            max_w, max_h = area_pixel_size(ws, area)

            pil = PILImage.open(photo_path)
            # PIL versiyonuna göre uyumlu thumbnail
            try:
                pil.thumbnail((max_w, max_h), PILImage.Resampling.LANCZOS)
            except AttributeError:
                # Eski PIL versiyonları için
                pil.thumbnail((max_w, max_h), PILImage.LANCZOS)

            tmp = os.path.join(temp_dir, f"_tmp_{i}.jpg")
            pil.save(tmp)
            temp_files.append(tmp)

            img = Image(tmp)
            img.anchor = smart_anchor(ws, area, pil.height)
            ws.add_image(img)

        # Print area ve sayfa ayarları
        setup_print_area(ws)

        # Geçici Excel dosyası oluştur
        temp_xlsx = os.path.join(temp_dir, f"rapor-{safe_date}-{uuid.uuid4().hex[:8]}.xlsx")
        wb.save(temp_xlsx)

        # PDF oluştur
        pdf_filename = f"rapor-{safe_date}-{uuid.uuid4().hex[:8]}.pdf"
        pdf_filepath = os.path.join(OUTPUT_DIR, pdf_filename)

        # Önce xlsx2pdf ile dene
        pdf_created = False
        error_messages = []
        
        print(f"PDF oluşturma başlıyor: {temp_xlsx} -> {pdf_filepath}")
        
        try:
            result = excel_to_pdf_xlsx2pdf(temp_xlsx, pdf_filepath)
            print(f"xlsx2pdf sonucu: {result}")
            if result:
                pdf_created = True
                print("PDF başarıyla oluşturuldu (xlsx2pdf)")
        except Exception as e:
            error_msg = f"xlsx2pdf hatası: {type(e).__name__}: {str(e)}"
            error_messages.append(error_msg)
            print(error_msg)
            import traceback
            traceback.print_exc()
        
        if not pdf_created:
            print("LibreOffice deneniyor...")
            try:
                result = excel_to_pdf_libreoffice(temp_xlsx, pdf_filepath)
                print(f"LibreOffice sonucu: {result}")
                if result:
                    pdf_created = True
                    print("PDF başarıyla oluşturuldu (LibreOffice)")
            except Exception as e:
                error_msg = f"LibreOffice hatası: {type(e).__name__}: {str(e)}"
                error_messages.append(error_msg)
                print(error_msg)
        
        if not pdf_created:
            error_msg = "PDF oluşturulamadı. "
            if error_messages:
                error_msg += " ".join(error_messages)
            else:
                error_msg += "xlsx2pdf veya LibreOffice gerekli."
            print(f"PDF oluşturma başarısız: {error_msg}")
            raise Exception(error_msg)

        return pdf_filepath

    finally:
        # Geçici dosyaları temizle
        for f in temp_files:
            if os.path.exists(f):
                try:
                    os.remove(f)
                except:
                    pass
        
        # Geçici dizini temizle
        try:
            shutil.rmtree(temp_dir)
        except:
            pass
